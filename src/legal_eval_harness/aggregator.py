from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .schemas import COARSE_ERROR_TAGS, DATA_ASSET_ROUTES, DATA_ROUTES, RESPONSE_POLICIES
from .utils import json_loads_or_none, parse_bool


def _iter_coarse_tags(values: pd.Series) -> list[str]:
    tags: list[str] = []
    for value in values:
        parsed = json_loads_or_none(value) or []
        for item in parsed:
            if isinstance(item, dict):
                tag = item.get("coarse_error_tag")
                if tag:
                    tags.append(str(tag))
            elif isinstance(item, str):
                tags.append(item)
    return tags


def _recommended_actions(top_tags: list[str]) -> str:
    actions = []
    if "missing_facts" in top_tags:
        actions.append("expand intake/SFT samples for missing-facts awareness")
    if "overclaim" in top_tags:
        actions.append("build preference pairs for overclaim control")
    if "missing_evidence_warning" in top_tags:
        actions.append("add evidence-risk warning exemplars")
    if "needs_human_review" in top_tags:
        actions.append("calibrate high-risk review queue")
    return "; ".join(actions) or "continue diagnostic eval expansion"


BADCASE_CARD_COLUMNS = [
    "sample_id",
    "source_dataset",
    "task_category",
    "run_id",
    "model_alias",
    "version",
    "main_error_type",
    "risk_level",
    "judge_confidence",
    "workflow_status",
    "response_policy",
    "data_asset_routes",
    "data_route",
    "priority",
    "route_reason",
    "route_subtype",
    "score_rate",
    "judge_reason",
    "output_preview",
]


def _build_badcase_cards(badcase_base: pd.DataFrame, *, limit: int = 80) -> pd.DataFrame:
    """Create review cards with route and task diversity."""
    cards = badcase_base.copy()
    for column in BADCASE_CARD_COLUMNS:
        if column not in cards.columns:
            cards[column] = ""
    cards["output_preview"] = cards["output_text"].astype(str).str.slice(0, 220)
    cards["_priority_rank"] = cards["priority"].map({"P0": 0, "P1": 1, "P2": 2}).fillna(9)
    cards["_route_rank"] = cards["response_policy"].map(
        {
            "block": 0,
            "human_review": 1,
            "clarify": 2,
            "grounded_answer": 3,
            "auto_answer": 4,
        }
    ).fillna(9)
    cards["_score_rate"] = pd.to_numeric(cards["score_rate"], errors="coerce").fillna(1.0)
    sort_cols = [
        "_route_rank",
        "task_category",
        "_priority_rank",
        "_score_rate",
        "sample_id",
        "version",
        "model_alias",
    ]
    cards = cards.sort_values(sort_cols)

    selected: list[int] = []
    selected_run_ids: set[str] = set()
    selected_card_keys: set[tuple[str, str, str]] = set()

    def card_key(row: pd.Series) -> tuple[str, str, str]:
        return (str(row["sample_id"]), str(row["version"]), str(row["response_policy"]))

    def add_first(group: pd.DataFrame, *, allow_duplicate_key: bool = False) -> None:
        for idx, row in group.iterrows():
            run_id = str(row["run_id"])
            key = card_key(row)
            if run_id not in selected_run_ids:
                if key in selected_card_keys and not allow_duplicate_key:
                    continue
                selected.append(idx)
                selected_run_ids.add(run_id)
                selected_card_keys.add(key)
                break

    task_order = ["consultation", "case_analysis", "document_drafting"]
    for response_policy in ["block", "human_review", "clarify", "grounded_answer", "auto_answer"]:
        route_group = cards[cards["response_policy"] == response_policy]
        for task_category in task_order:
            add_first(route_group[route_group["task_category"] == task_category])

    for _, group in cards.groupby(["response_policy", "main_error_type"], sort=False):
        add_first(group)

    for idx, row in cards.iterrows():
        if len(selected) >= limit:
            break
        run_id = str(row["run_id"])
        key = card_key(row)
        if run_id not in selected_run_ids and key not in selected_card_keys:
            selected.append(idx)
            selected_run_ids.add(run_id)
            selected_card_keys.add(key)

    for idx, row in cards.iterrows():
        if len(selected) >= limit:
            break
        run_id = str(row["run_id"])
        if run_id not in selected_run_ids:
            selected.append(idx)
            selected_run_ids.add(run_id)
            selected_card_keys.add(card_key(row))

    return cards.loc[selected, BADCASE_CARD_COLUMNS].reset_index(drop=True)


def _numeric_column(frame: pd.DataFrame, column: str) -> pd.Series:
    if column not in frame.columns:
        return pd.Series([0] * len(frame), index=frame.index, dtype="float64")
    return pd.to_numeric(frame[column], errors="coerce").fillna(0)


def _canonical_routing(routing: pd.DataFrame) -> pd.DataFrame:
    """Add canonical routing fields while accepting legacy CSV artifacts."""
    result = routing.copy()
    if "data_route" not in result.columns:
        result["data_route"] = "eval"
    if "response_policy" not in result.columns:
        legacy_decision = result.get("release_decision", pd.Series("", index=result.index)).fillna("")
        result["response_policy"] = legacy_decision.where(
            legacy_decision.isin(RESPONSE_POLICIES),
            result["data_route"].map(lambda value: "human_review" if value == "human_review" else "auto_answer"),
        )
    if "release_decision" not in result.columns:
        result["release_decision"] = result["response_policy"]
    if "workflow_status" not in result.columns:
        result["workflow_status"] = result["response_policy"].map(
            {"block": "blocked", "human_review": "pending_review"}
        ).fillna("released")
    if "data_asset_routes" not in result.columns:
        result["data_asset_routes"] = result["data_route"].map(
            lambda route: [route] if route in DATA_ASSET_ROUTES else []
        )
    return result


def build_executive_dashboard(
    *,
    runs: pd.DataFrame,
    scores: pd.DataFrame,
    routing: pd.DataFrame,
    output_path: str | Path,
) -> dict[str, Any]:
    scores = scores.copy()
    scores["score_rate"] = scores["score_rate"].astype(float)
    if "needs_human_review" in scores.columns:
        scores["needs_human_review"] = scores["needs_human_review"].map(parse_bool)
    runs = runs.copy()
    routing = _canonical_routing(routing)
    for frame in [scores, runs, routing]:
        if "source_dataset" not in frame.columns:
            frame["source_dataset"] = "unknown"
        if "task_category" not in frame.columns:
            frame["task_category"] = "unknown"

    pivot = scores.pivot_table(
        index=["sample_id", "model_alias"],
        columns="version",
        values="score_rate",
        aggfunc="mean",
    )
    if {"V0", "V3"}.issubset(set(pivot.columns)):
        deltas = pivot["V3"] - pivot["V0"]
        avg_delta = round(float(deltas.mean()), 3)
    else:
        avg_delta = 0.0

    all_tags = _iter_coarse_tags(scores["error_tags"])
    tag_counts = Counter(all_tags)
    top_3 = [tag for tag, _ in tag_counts.most_common(3)]
    if "has_nonempty_output" in runs.columns:
        has_nonempty_output = runs["has_nonempty_output"].map(parse_bool)
    else:
        has_nonempty_output = runs.get("output_text", pd.Series("", index=runs.index)).map(
            lambda value: bool(str(value).strip()) if not pd.isna(value) else False
        )
    dashboard = {
        "total_samples": int(scores["sample_id"].nunique()),
        "total_api_run_rows": int(len(runs)),
        # Backward-compatible metric name. This counts API run rows, not non-empty answers.
        "total_runs": int(len(runs)),
        "nonempty_answer_count": int(has_nonempty_output.sum()),
        "empty_answer_count": int((~has_nonempty_output).sum()),
        "total_task_categories": int(scores["task_category"].nunique()) if "task_category" in scores.columns else 0,
        "total_source_datasets": int(scores["source_dataset"].nunique()) if "source_dataset" in scores.columns else 0,
        "avg_v0_score": round(float(scores.loc[scores["version"] == "V0", "score_rate"].mean()), 3),
        "avg_v3_score": round(float(scores.loc[scores["version"] == "V3", "score_rate"].mean()), 3),
        "avg_score_delta": avg_delta,
        "high_risk_rate": round(float((scores["risk_level"] == "high").mean()), 3),
        "human_review_queue_size": int((routing["response_policy"] == "human_review").sum()),
        "blocked_response_count": int((routing["response_policy"] == "block").sum()),
        "avg_latency_ms": round(float(_numeric_column(runs, "latency_ms").mean()), 1),
        "total_estimated_cost": round(float(_numeric_column(runs, "estimated_cost").sum()), 6),
        "top_3_error_tags": ", ".join(top_3),
        "recommended_data_actions": _recommended_actions(top_3),
        "dashboard_boundary": "Data production decision panel; diagnostic comparison only; not a model leaderboard.",
    }

    sample_cols = ["sample_id"]
    for col in ["source_dataset", "task_category"]:
        if col in runs.columns:
            sample_cols.append(col)
    sample_catalog = runs[sample_cols].drop_duplicates()
    dataset_coverage = (
        sample_catalog.groupby(["source_dataset", "task_category"], as_index=False)
        .agg(samples=("sample_id", "nunique"))
        .merge(
            runs.groupby(["source_dataset", "task_category"], as_index=False).agg(runs=("run_id", "count")),
            on=["source_dataset", "task_category"],
            how="left",
        )
    )
    dataset_coverage["coverage_note"] = "Core samples emphasize quality; extended samples exercise scale and task coverage."

    task_tag_rows = []
    for task_category, grp in scores.groupby("task_category"):
        tags = Counter(_iter_coarse_tags(grp["error_tags"]))
        routes = routing[routing["task_category"] == task_category]["response_policy"].value_counts().to_dict()
        response_policy_mix = "; ".join(
            f"{response_policy}:{count}" for response_policy, count in sorted(routes.items())
        )
        task_tag_rows.append(
            {
                "task_category": task_category,
                "samples": int(grp["sample_id"].nunique()),
                "runs": int(len(grp)),
                "avg_score_rate": round(float(grp["score_rate"].mean()), 3),
                "high_risk_rate": round(float((grp["risk_level"] == "high").mean()), 3),
                "human_review_rate": round(float(grp["needs_human_review"].mean()), 3),
                "top_error_tags": ", ".join([tag for tag, _ in tags.most_common(3)]),
                "response_policy_mix": response_policy_mix,
                # Backward-compatible column name.
                "route_mix": response_policy_mix,
                "data_action": _recommended_actions([tag for tag, _ in tags.most_common(3)]),
            }
        )
    task_category_summary = pd.DataFrame(task_tag_rows)

    version_summary = (
        scores.groupby("version", as_index=False)
        .agg(
            avg_score_rate=("score_rate", "mean"),
            high_risk_rate=("risk_level", lambda x: (x == "high").mean()),
            human_review_rate=("needs_human_review", "mean"),
        )
        .round(3)
    )
    if "workflow_condition" in runs.columns:
        workflow_lookup = runs[["version", "workflow_condition", "workflow_name"]].drop_duplicates()
        version_summary = version_summary.merge(workflow_lookup, on="version", how="left")

    operational_summary = runs.copy()
    for col in ["latency_ms", "input_tokens", "output_tokens", "total_tokens", "estimated_cost"]:
        if col not in operational_summary.columns:
            operational_summary[col] = 0
        operational_summary[col] = pd.to_numeric(operational_summary[col], errors="coerce").fillna(0)
    group_cols = [
        col
        for col in ["task_category", "model_alias", "version", "workflow_condition", "workflow_name"]
        if col in operational_summary.columns
    ]
    cost_latency_summary = (
        operational_summary.groupby(group_cols, as_index=False)
        .agg(
            runs=("run_id", "count"),
            avg_latency_ms=("latency_ms", "mean"),
            avg_input_tokens=("input_tokens", "mean"),
            avg_output_tokens=("output_tokens", "mean"),
            total_estimated_cost=("estimated_cost", "sum"),
        )
        .round(4)
    )

    policy_base = scores.merge(
        routing[["run_id", "workflow_status", "response_policy", "data_asset_routes", "data_route", "priority"]],
        on="run_id",
        how="left",
    )
    if "workflow_condition" in runs.columns:
        policy_base = policy_base.merge(
            runs[["run_id", "workflow_condition", "workflow_name", "latency_ms", "estimated_cost"]],
            on="run_id",
            how="left",
        )
    else:
        policy_base["workflow_condition"] = policy_base["version"]
        policy_base["workflow_name"] = policy_base["version"]
        policy_base["latency_ms"] = 0
        policy_base["estimated_cost"] = 0
    deployment_policy = (
        policy_base.groupby(["task_category", "workflow_condition", "workflow_name"], as_index=False)
        .agg(
            runs=("run_id", "count"),
            avg_score_rate=("score_rate", "mean"),
            high_risk_rate=("risk_level", lambda x: (x == "high").mean()),
            human_review_rate=("response_policy", lambda x: (x == "human_review").mean()),
            blocked_response_rate=("response_policy", lambda x: (x == "block").mean()),
            avg_latency_ms=("latency_ms", "mean"),
            total_estimated_cost=("estimated_cost", "sum"),
        )
        .round(4)
    )
    deployment_policy["policy_hint"] = deployment_policy.apply(
        lambda row: (
            "candidate for auto-answer only if critical checks pass"
            if row["high_risk_rate"] <= 0.05 and row["human_review_rate"] <= 0.15
            else "route to human review or stronger workflow before release"
            if row["high_risk_rate"] >= 0.2
            else "limited release with monitoring and targeted data production"
        ),
        axis=1,
    )

    route_summary = (
        routing.groupby(["data_route", "task_category"], as_index=False)
        .agg(count=("run_id", "count"), example_sample_ids=("sample_id", lambda x: ", ".join(sorted(set(x))[:5])))
        .sort_values("count", ascending=False)
    )

    workflow_release_summary = (
        routing.groupby(["workflow_status", "response_policy"], as_index=False)
        .agg(count=("run_id", "count"))
        .sort_values("count", ascending=False)
    )

    asset_rows: list[dict[str, str]] = []
    if "data_asset_routes" in routing.columns:
        for _, route_row in routing.iterrows():
            assets = json_loads_or_none(route_row.get("data_asset_routes")) or []
            for asset in assets:
                asset_rows.append(
                    {
                        "run_id": str(route_row.get("run_id", "")),
                        "task_category": str(route_row.get("task_category", "")),
                        "data_asset_route": str(asset),
                    }
                )
    asset_route_summary = (
        pd.DataFrame(asset_rows)
        .groupby(["data_asset_route", "task_category"], as_index=False)
        .agg(count=("run_id", "count"))
        .sort_values("count", ascending=False)
        if asset_rows
        else pd.DataFrame(columns=["data_asset_route", "task_category", "count"])
    )

    model_patterns = []
    for model_alias, grp in scores.groupby("model_alias"):
        tags = Counter(_iter_coarse_tags(grp["error_tags"]))
        model_patterns.append(
            {
                "model_alias": model_alias,
                "dominant_error_patterns": ", ".join([tag for tag, _ in tags.most_common(3)]),
                "workflow_delta_note": "Diagnostic comparison only; use this for workflow effects and data needs, not model ranking.",
                "human_review_rate": round(float(grp["needs_human_review"].mean()), 3),
            }
        )
    model_patterns_df = pd.DataFrame(model_patterns)

    badcase_base = routing.merge(
        scores.assign(
            total_score=scores["total_score"] if "total_score" in scores.columns else "",
            max_score=scores["max_score"] if "max_score" in scores.columns else "",
            judge_reason=scores["judge_reason"] if "judge_reason" in scores.columns else "",
            judge_confidence=scores["judge_confidence"] if "judge_confidence" in scores.columns else "",
        )[
            [
                "run_id",
                "total_score",
                "max_score",
                "score_rate",
                "judge_reason",
                "judge_confidence",
                "needs_human_review",
            ]
        ],
        on="run_id",
        how="left",
    ).merge(
        runs[["run_id", "output_text", "output_length"]],
        on="run_id",
        how="left",
    )
    badcase_cards = _build_badcase_cards(badcase_base, limit=80)

    taxonomy_df = pd.DataFrame(
        {
            "coarse_error_tag": COARSE_ERROR_TAGS,
            "definition": [
                "Answer misses necessary facts before analysis.",
                "Answer states a stronger conclusion than facts support.",
                "Answer fails to warn that evidence is required.",
                "Answer relies on a legal basis that is not verified.",
                "Answer fabricates laws, cases, institutions, or citations.",
                "Answer weakly connects facts to legal rules.",
                "Answer misses timing, procedure, limitation, or channel warnings.",
                "Answer ignores jurisdiction or local-practice uncertainty.",
                "Answer suggests potentially unsafe action.",
                "Sample should be calibrated by human review.",
            ],
        }
    )
    route_taxonomy_df = pd.DataFrame(
        {
            "data_route": DATA_ROUTES,
            "meaning": [
                "held-out diagnostic evaluation",
                "supervised fine-tuning candidate",
                "preference pair candidate",
                "regression badcase set",
                "repeat-check regression asset",
                "human calibration queue",
            ],
        }
    )
    response_policy_taxonomy_df = pd.DataFrame(
        {
            "response_policy": RESPONSE_POLICIES,
            "meaning": [
                "answer without retrieval after release checks pass",
                "answer only with retrieval/citation grounding",
                "ask for missing material facts before analysis",
                "send to human review before release",
                "do not release the response",
            ],
        }
    )

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame([dashboard]).to_excel(writer, sheet_name="Executive_Dashboard", index=False)
        dataset_coverage.to_excel(writer, sheet_name="Dataset_Coverage", index=False)
        task_category_summary.to_excel(writer, sheet_name="Task_Category_Summary", index=False)
        badcase_cards.to_excel(writer, sheet_name="Badcase_Cards", index=False)
        version_summary.to_excel(writer, sheet_name="Version_Summary", index=False)
        cost_latency_summary.to_excel(writer, sheet_name="Cost_Latency", index=False)
        deployment_policy.to_excel(writer, sheet_name="Deployment_Policy", index=False)
        route_summary.to_excel(writer, sheet_name="Data_Routing_Summary", index=False)
        workflow_release_summary.to_excel(writer, sheet_name="Workflow_Release", index=False)
        asset_route_summary.to_excel(writer, sheet_name="Data_Asset_Routes", index=False)
        model_patterns_df.to_excel(writer, sheet_name="Model_Error_Patterns", index=False)
        taxonomy_df.to_excel(writer, sheet_name="Error_Taxonomy", index=False)
        route_taxonomy_df.to_excel(writer, sheet_name="Data_Route_Taxonomy", index=False)
        response_policy_taxonomy_df.to_excel(writer, sheet_name="Response_Policy", index=False)
        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for ws in workbook.worksheets:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for column_cells in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in column_cells[:80])
                width = min(max(max_len + 2, 12), 42)
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
    return dashboard
