#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import importlib.util
import json
import subprocess
from pathlib import Path
from typing import Iterable

try:
    import yaml
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit(
        "FAIL: project dependencies are required. Install with "
        '`python3 -m pip install -e ".[test]"`.'
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
EMPTY_SHA256 = "e3b0c44298fc1c149afbf4c8996fb92427ae41e4649b934ca495991b7852b855"
CANONICAL_DATA_ASSETS = {"eval", "sft", "preference", "badcase", "regression"}
CANONICAL_TRACE_RECOMMENDATIONS = {
    "blocked",
    "human_review_required",
    "candidate_limited_release",
}


def fail(msg: str) -> None:
    raise SystemExit(f"FAIL: {msg}")


def _relative(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def read_jsonl(path: Path) -> list[dict]:
    rows: list[dict] = []
    with path.open(encoding="utf-8") as handle:
        for line_no, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            try:
                rows.append(json.loads(line))
            except Exception as exc:
                fail(f"JSONL parse error in {_relative(path)}:{line_no}: {exc}")
    return rows


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    try:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    except (UnicodeDecodeError, csv.Error) as exc:
        fail(f"CSV parse error in {_relative(path)}: {exc}")


def iter_tracked_output_files(suffix: str) -> Iterable[Path]:
    """Ignore local/raw artifacts and validate only the public Git evidence surface."""
    try:
        result = subprocess.run(
            ["git", "ls-files", "outputs"],
            cwd=ROOT,
            check=True,
            capture_output=True,
            text=True,
        )
    except (OSError, subprocess.CalledProcessError) as exc:
        fail(f"Could not enumerate tracked public outputs: {exc}")
    for rel in result.stdout.splitlines():
        path = ROOT / rel
        if path.exists() and path.suffix == suffix:
            yield path


def iter_yaml_files() -> Iterable[Path]:
    yield from ROOT.glob("*.yaml")
    yield from (ROOT / "configs").rglob("*.yaml")
    yield from iter_tracked_output_files(".yaml")
    # Newly added public manifests may not be in the Git index yet.
    for rel in ["outputs/product_boundary_pilot_mock/artifact_manifest.yaml"]:
        path = ROOT / rel
        if path.exists():
            yield path


def check_yaml() -> None:
    for path in iter_yaml_files():
        try:
            yaml.safe_load(path.read_text(encoding="utf-8"))
        except Exception as exc:
            fail(f"YAML parse error in {_relative(path)}: {exc}")


def check_jsonl() -> None:
    for path in (ROOT / "data").rglob("*.jsonl"):
        read_jsonl(path)
    for path in iter_tracked_output_files(".jsonl"):
        read_jsonl(path)


def check_csv() -> None:
    paths = list((ROOT / "data").rglob("*.csv")) + list(iter_tracked_output_files(".csv"))
    for path in paths:
        try:
            with path.open(encoding="utf-8-sig", newline="") as handle:
                reader = csv.reader(handle)
                header = next(reader, None)
                if not header:
                    fail(f"Empty CSV file: {_relative(path)}")
                expected_width = len(header)
                for line_no, row in enumerate(reader, start=2):
                    if len(row) != expected_width:
                        fail(
                            f"CSV width mismatch in {_relative(path)}:{line_no}; "
                            f"expected {expected_width}, found {len(row)}"
                        )
        except UnicodeDecodeError:
            fail(f"CSV is not UTF-8 decodable: {_relative(path)}")
        except csv.Error as exc:
            fail(f"CSV parse error in {_relative(path)}: {exc}")


def check_case_bank() -> None:
    path = ROOT / "data/eval_sets/legal_product_boundary_pilot_v1.jsonl"
    rows = read_jsonl(path)
    if len(rows) != 50:
        fail(f"Expected 50 product-boundary cases, found {len(rows)}")
    ids = [row.get("case_id") for row in rows]
    if any(not case_id for case_id in ids):
        fail("Missing case_id in product-boundary case bank")
    if len(ids) != len(set(ids)):
        fail("Duplicate case_id in product-boundary case bank")
    for row in rows:
        if "expected_data_route" in row:
            fail(f"Legacy expected_data_route remains in {row.get('case_id')}")
        routes = row.get("expected_data_asset_routes")
        if not isinstance(routes, list) or not routes:
            fail(f"Missing expected_data_asset_routes in {row.get('case_id')}")
        if invalid := set(map(str, routes)) - CANONICAL_DATA_ASSETS:
            fail(f"Invalid data asset routes in {row.get('case_id')}: {sorted(invalid)}")


def check_a5_cases() -> None:
    path = ROOT / "data/eval_sets/legal_agent_multiturn_intake_pilot_v1.jsonl"
    rows = read_jsonl(path)
    allowed_profiles = {"cooperative", "dependent", "withdrawn", "adversarial"}
    for row in rows:
        case_id = row.get("case_id")
        raw_profile = row.get("simulator_profile") or row.get("user_behavior")
        profile = str(raw_profile or "").removesuffix("_client")
        if profile not in allowed_profiles:
            fail(f"Invalid simulator profile/user behavior in {case_id}: {raw_profile}")
        if not row.get("material_facts_to_elicit"):
            fail(f"Missing material_facts_to_elicit in {case_id}")
        if "expected_release_policy" in row or "expected_data_route" in row:
            fail(f"Legacy A5 route field remains in {case_id}")
        recommendation = row.get("expected_trace_review_recommendation")
        if not recommendation:
            fail(f"Missing expected_trace_review_recommendation in {case_id}")
        elif recommendation not in CANONICAL_TRACE_RECOMMENDATIONS:
            fail(f"Invalid expected_trace_review_recommendation in {case_id}: {recommendation}")
        routes = row.get("expected_data_asset_routes")
        if not isinstance(routes, list) or not routes:
            fail(f"Missing expected_data_asset_routes in {case_id}")
        if invalid := set(map(str, routes)) - CANONICAL_DATA_ASSETS:
            fail(f"Invalid A5 data asset routes in {case_id}: {sorted(invalid)}")


def check_focused_config() -> None:
    path = ROOT / "configs/experiments/legal_agent_product_eval_v2_focused.yaml"
    data = yaml.safe_load(path.read_text(encoding="utf-8"))
    experiment = data.get("experiment", {})
    scope = data.get("scope", {})
    if experiment.get("status") != "planned":
        fail("Focused V2 config must remain status: planned")
    if int(scope.get("expected_api_run_records", 0)) != 450:
        fail("Focused V2 config must describe 450 planned API run records")


def _metric_map(path: Path) -> dict[str, str]:
    rows = read_csv_rows(path)
    return {str(row.get("metric", "")): str(row.get("value", "")) for row in rows}


def _expect_metric(metrics: dict[str, str], key: str, expected: str) -> None:
    actual = metrics.get(key)
    if actual != expected:
        fail(f"Metric {key!r} expected {expected!r}, found {actual!r}")


def check_api_evidence() -> None:
    base = ROOT / "outputs/product_boundary_api_pilot_v1"
    required = [
        "README.md",
        "artifact_manifest.yaml",
        "metrics_summary.csv",
        "release_gate_summary.csv",
        "human_calibration_summary_priority_80.csv",
        "claim_entailment_summary.csv",
        "redacted_sample_outputs_20.csv",
    ]
    for filename in required:
        if not (base / filename).exists():
            fail(f"Missing API evidence artifact: {_relative(base / filename)}")

    manifest = yaml.safe_load((base / "artifact_manifest.yaml").read_text(encoding="utf-8"))
    if str(manifest.get("schema_version")) != "2.0":
        fail("API evidence manifest must use schema_version 2.0")
    source_run = manifest.get("source_run") or {}
    expected_source = {
        "api_runs": 300,
        "nonempty_answers": 271,
        "empty_answers": 29,
        "cases": 12,
        "models": 5,
        "workflows": 5,
    }
    for key, expected in expected_source.items():
        if int(source_run.get(key, -1)) != expected:
            fail(f"API manifest {key} expected {expected}, found {source_run.get(key)!r}")

    metrics = _metric_map(base / "metrics_summary.csv")
    for key, expected in {
        "api_run_count": "300",
        "nonempty_answer_count": "271",
        "empty_answer_count": "29",
        "unique_cases": "12",
        "models": "5",
        "workflow_versions": "5",
        "reviewable_claim_count": "716",
        "reviewable_claim_strict_citation_defect_flag_count": "619",
        "reviewable_claim_needs_review_count": "660",
        "all_claim_source_boundary_blocker_count": "60",
        "formal_reviewer_iaa": "not_reported",
        "formal_judge_human_agreement": "not_reported",
    }.items():
        _expect_metric(metrics, key, expected)
    forbidden_metrics = {"real_model_outputs", "human_calibration_judge_human_agreement_rate"}
    present = forbidden_metrics.intersection(metrics)
    if present:
        fail(f"Legacy or misleading API evidence metrics remain: {sorted(present)}")

    review = _metric_map(base / "human_calibration_summary_priority_80.csv")
    _expect_metric(review, "priority_review_row_count", "80")
    _expect_metric(review, "rows_with_nonempty_answer", "51")
    _expect_metric(review, "rows_with_empty_answer", "29")
    _expect_metric(review, "completed_final_review_rows", "80")
    _expect_metric(review, "formal_reviewer_iaa", "not_reported")
    _expect_metric(review, "formal_judge_human_agreement", "not_reported")

    samples = read_csv_rows(base / "redacted_sample_outputs_20.csv")
    if len(samples) != 20:
        fail(f"Expected 20 redacted non-empty API samples, found {len(samples)}")
    for row in samples:
        if row.get("content_status") != "nonempty":
            fail("Redacted public API samples must exclude empty answers")
        if row.get("output_sha256") == EMPTY_SHA256:
            fail("Redacted public API sample contains the empty-output hash")
    sample_ids = {row.get("sample_id") for row in samples}
    if not {"LPB-RISK-001", "LPB-CITE-001", "LPB-CITE-002"}.issubset(sample_ids):
        fail("Redacted public API samples must include the flagship risk and citation cases")
    required_signal_columns = {
        "response_policy",
        "workflow_status",
        "main_error_type",
        "data_asset_routes",
    }
    if samples and not required_signal_columns.issubset(samples[0]):
        fail("Redacted public API samples are missing product decision signals")


def check_rag_evidence() -> None:
    base = ROOT / "outputs/rag_v2_focused_pilot_v1"
    metrics_path = base / "metrics_summary.csv"
    manifest_path = base / "artifact_manifest.yaml"
    if not metrics_path.exists() or not manifest_path.exists():
        fail("Missing focused RAG evidence package")
    metrics = _metric_map(metrics_path)
    for key, expected in {
        "api_run_records": "72",
        "reviewable_claim_rows": "630",
        "reviewable_claim_strict_citation_defect_rows": "555",
        "reviewable_claim_needs_review_rows": "591",
        "all_claim_source_boundary_blocker_rows": "75",
    }.items():
        _expect_metric(metrics, key, expected)
    for forbidden in {
        "model_outputs",
        "reviewable_claim_citation_issue_rows",
        "reviewable_claim_citation_issue_rate",
        "citation_gate_issue_rows",
        "citation_gate_issue_rate",
        "claim_release_blocker_rows",
        "claim_release_blocker_rate",
    }:
        if forbidden in metrics:
            fail(f"Ambiguous legacy RAG metric must not be public: {forbidden}")
    _expect_metric(metrics, "reviewable_claim_strict_citation_defect_rate", "0.881")
    _expect_metric(metrics, "reviewable_claim_needs_review_rate", "0.9381")
    _expect_metric(metrics, "all_claim_source_boundary_blocker_rate", "0.0425")
    manifest = yaml.safe_load(manifest_path.read_text(encoding="utf-8"))
    if str(manifest.get("schema_version")) != "2.0":
        fail("RAG evidence manifest must use schema_version 2.0")
    model_rows = read_csv_rows(base / "model_workflow_summary.csv")
    allowed_gate_decisions = {"blocked", "limited_release", "candidate_auto_answer"}
    if any(row.get("release_gate_decision") not in allowed_gate_decisions for row in model_rows):
        fail("RAG model-workflow release_gate_decision contains a non-canonical value")
    forbidden_columns = {
        "recommended_policy",
        "reviewable_citation_issue_count",
        "citation_gate_issue_count",
        "claim_release_blocker_count",
    }
    for filename in ["workflow_comparison.csv", "model_workflow_summary.csv", "redacted_sample_outputs_20.csv"]:
        rows = read_csv_rows(base / filename)
        if rows and forbidden_columns.intersection(rows[0]):
            fail(f"Ambiguous legacy RAG columns remain in {filename}")


def check_a5_evidence() -> None:
    base = ROOT / "outputs/a5_multiturn_intake_pilot_v1"
    metrics_path = base / "trace_metrics_summary.csv"
    manifest_path = base / "artifact_manifest.yaml"
    if not metrics_path.exists() or not manifest_path.exists():
        fail("Missing A5 evidence package")
    metrics = _metric_map(metrics_path)
    _expect_metric(metrics, "traces", "24")
    _expect_metric(metrics, "turns", "72")
    for forbidden in ["trace_pass_rate", "model_pass_rate", "behavior_pass_rate"]:
        if forbidden in metrics:
            fail(f"Uncalibrated A5 pass metric must not be published: {forbidden}")
    manifest = yaml.safe_load(manifest_path.read_text(encoding="utf-8"))
    if str(manifest.get("schema_version")) != "2.0":
        fail("A5 evidence manifest must use schema_version 2.0")
    if manifest.get("run_stage") != "pilot" or manifest.get("current_portfolio_evidence") is not True:
        fail("Primary A5 evidence must be marked as the current pilot")
    template_rows = read_csv_rows(base / "human_trace_calibration_template.csv")
    if not template_rows:
        fail("A5 human calibration template is empty")
    columns = set(template_rows[0])
    if "human_data_route" in columns or "human_data_asset_routes" not in columns:
        fail("A5 human calibration template must use canonical multi-label asset routes")
    if not {"human_response_policy", "human_workflow_status"}.issubset(columns):
        fail("A5 human calibration template must separate response policy and workflow status")


def check_mock_publication_boundary() -> None:
    base = ROOT / "outputs/product_boundary_pilot_mock"
    allowed = {"README.md", "artifact_manifest.yaml"}
    unexpected = sorted(path.name for path in base.iterdir() if path.is_file() and path.name not in allowed)
    if unexpected:
        fail(f"Synthetic mock artifacts must remain local/ignored; found {unexpected}")
    manifest = yaml.safe_load((base / "artifact_manifest.yaml").read_text(encoding="utf-8"))
    if manifest.get("artifact_class") != "SYNTHETIC_FIXTURE":
        fail("Mock artifact manifest must be marked SYNTHETIC_FIXTURE")


def check_source_provenance() -> None:
    path = ROOT / "data/rag_corpus/legal_sources.csv"
    rows = read_csv_rows(path)
    required = {
        "source_id",
        "provenance_status",
        "document_identifier",
        "source_version",
        "retrieved_at",
        "content_sha256",
        "license_or_origin",
        "publishable_as_authoritative_source",
    }
    columns = set(rows[0]) if rows else set()
    if missing := required - columns:
        fail(f"RAG source provenance columns missing: {sorted(missing)}")
    ids = [row["source_id"] for row in rows]
    if len(ids) != len(set(ids)):
        fail("RAG source_id values must be unique")
    for row in rows:
        if len(row.get("content_sha256", "")) != 64:
            fail(f"Invalid content hash for source {row.get('source_id')}")
        if row.get("source_url") != "self_authored":
            if row.get("provenance_status") != "summary_requires_primary_source_verification":
                fail(f"External summary lacks verification boundary: {row.get('source_id')}")
            if row.get("publishable_as_authoritative_source", "").lower() != "false":
                fail(f"Unverified source marked authoritative: {row.get('source_id')}")

    practice_manifest = yaml.safe_load(
        (ROOT / "data/practice_benchmark_pilot/dataset_manifest.yaml").read_text(encoding="utf-8")
    )
    practice_source = (practice_manifest.get("sources") or [{}])[0]
    if "verify" not in str(practice_source.get("license_status", "")).lower():
        fail("External practice benchmark must disclose unverified license metadata")


def check_generated_output_schema(output_dir: Path) -> None:
    routing_path = output_dir / "data_routing.csv"
    workbook_path = output_dir / "executive_dashboard.xlsx"
    if not routing_path.exists() or not workbook_path.exists():
        fail(f"Generated mock output is incomplete: {_relative(output_dir)}")
    rows = read_csv_rows(routing_path)
    columns = set(rows[0]) if rows else set()
    required_columns = {
        "workflow_status",
        "response_policy",
        "data_asset_routes",
        "candidate_for_reuse",
        "requires_correction",
        "gold_approved",
    }
    if missing := required_columns - columns:
        fail(f"Generated routing schema missing columns: {sorted(missing)}")
    for row in rows:
        try:
            assets = json.loads(row["data_asset_routes"])
        except json.JSONDecodeError as exc:
            fail(f"Invalid data_asset_routes JSON for run {row.get('run_id')}: {exc}")
        if not isinstance(assets, list):
            fail(f"data_asset_routes must serialize a list for run {row.get('run_id')}")
        if row.get("gold_approved", "").lower() == "true" and row.get("requires_correction", "").lower() == "true":
            fail(f"Uncorrected routing row marked gold: {row.get('run_id')}")

    workbook = load_workbook(workbook_path, read_only=True)
    required_sheets = {"Workflow_Release", "Data_Asset_Routes", "Response_Policy"}
    if missing := required_sheets - set(workbook.sheetnames):
        fail(f"Generated dashboard missing sheets: {sorted(missing)}")


def check_ci() -> None:
    path = ROOT / ".github/workflows/ci.yml"
    if not path.exists():
        fail("Missing GitHub Actions CI workflow")
    text = path.read_text(encoding="utf-8")
    for token in ['"3.11"', '"3.12"', "pip install -e", "--generated-output-dir"]:
        if token not in text:
            fail(f"CI workflow missing required validation token: {token}")


def check_source_install() -> None:
    spec = importlib.util.find_spec("legal_eval_harness")
    if spec is None or not spec.submodule_search_locations:
        fail('legal_eval_harness is not importable; install with `pip install -e ".[test]"`')
    expected = (ROOT / "src/legal_eval_harness").resolve()
    locations = {Path(path).resolve() for path in spec.submodule_search_locations}
    if expected not in locations:
        fail(
            "Python resolves legal_eval_harness outside the current src tree; "
            'remove the stale install and run `pip install -e ".[test]"`'
        )


def check_markdown() -> None:
    targets = [
        "README.md",
        "docs/case_study.md",
        "docs/deepseek_product_note.md",
        "docs/human_review_methodology.md",
        "docs/role_fit_deepseek_data_pm.md",
        "docs/final_portfolio_findings.md",
        "docs/results_product_boundary_eval.md",
        "docs/model_boundary_memo.md",
        "docs/a5_multiturn_pilot_results.md",
        "docs/rag_v2_focused_results.md",
    ]
    for rel in targets:
        path = ROOT / rel
        if not path.exists():
            fail(f"Missing portfolio document: {rel}")
        lines = path.read_text(encoding="utf-8").splitlines()
        if len(lines) < 12:
            fail(f"Portfolio document may be truncated: {rel}")
        long_lines = [
            line_no
            for line_no, line in enumerate(lines, start=1)
            if len(line) > 350 and not line.lstrip().startswith("|")
        ]
        if long_lines:
            fail(f"Markdown has very long non-table lines in {rel}: {long_lines[:5]}")


def check_claim_boundaries() -> None:
    readme = (ROOT / "README.md").read_text(encoding="utf-8")
    lowered = readme.lower()
    for token in ["271", "29", "证据边界", "human_review_methodology.md", "deepseek_product_note.md"]:
        if token.lower() not in lowered:
            fail(f"README is missing required evidence-boundary token: {token}")

    active_paths = [ROOT / "README.md", ROOT / "assets/product_eval_system_preview.svg"]
    active_paths.extend(path for path in (ROOT / "docs").glob("*.md"))
    active_paths.extend(
        [
            ROOT / "outputs/product_boundary_api_pilot_v1/README.md",
            ROOT / "outputs/a5_multiturn_intake_pilot_v1/README.md",
            ROOT / "outputs/rag_v2_focused_pilot_v1/README.md",
        ]
    )
    joined = "\n".join(path.read_text(encoding="utf-8").lower() for path in active_paths if path.exists())
    forbidden_phrases = [
        "92.5%",
        "7/8",
        "75.0%",
        "300 条真实输出",
        "300 real model outputs",
        "licensed_practice_benchmark_adapted",
        "production-ready legal advice",
        "production ready legal advice",
        "planned 450-output focused experiment 已完成",
    ]
    for phrase in forbidden_phrases:
        if phrase in joined:
            fail(f"Unsupported or stale public claim remains: {phrase}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate portfolio evidence, boundaries, and generated schemas")
    parser.add_argument("--generated-output-dir", type=Path, default=None)
    args = parser.parse_args()

    check_yaml()
    check_jsonl()
    check_csv()
    check_case_bank()
    check_a5_cases()
    check_focused_config()
    check_api_evidence()
    check_rag_evidence()
    check_a5_evidence()
    check_mock_publication_boundary()
    check_source_provenance()
    check_ci()
    check_source_install()
    check_markdown()
    check_claim_boundaries()
    if args.generated_output_dir:
        check_generated_output_schema(args.generated_output_dir.resolve())
    print("Portfolio release validation passed.")


if __name__ == "__main__":
    main()
